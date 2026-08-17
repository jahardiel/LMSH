import pandas as pd
import openpyxl
import numpy as np
import io
from typing import Dict, List, Tuple

def crear_plantilla_excel(filepath: str = "plantilla_medicion_concreto.xlsx"):
    """
    Crea una plantilla de Excel (.xlsx) estructurada de forma limpia:
    - Hoja 1: Lecturas_Campo (Solo las lecturas de temperatura del concreto)
    - Hoja 2: Parametros_Instrumento (Resolución, deriva y homogeneidad)
    - Hoja 3: Calibracion_Termometro (Tabla de calibración)
    """
    wb = openpyxl.Workbook()

    # Hoja 1: Lecturas_Campo
    ws1 = wb.active
    ws1.title = "Lecturas_Campo"
    ws1.append(["Lectura_No", "Temperatura_Concreto_C"])

    lecturas_demo = [23.40, 23.60, 23.50, 23.70, 23.50, 23.40]
    for i, val in enumerate(lecturas_demo, 1):
        ws1.append([i, val])

    # Hoja 2: Parametros_Instrumento
    ws2 = wb.create_sheet(title="Parametros_Instrumento")
    ws2.append(["Parametro", "Valor", "Unidad", "Descripción"])
    ws2.append(["Resolucion", 0.01, "°C", "Resolución del indicador del termómetro"])
    ws2.append(["Deriva_Estimada", 0.05, "°C", "Deriva instrumental máxima esperada"])
    ws2.append(["Homogeneidad_Concreto", 0.10, "°C", "Variabilidad o gradiente térmico de la mezcla"])

    # Hoja 3: Calibracion_Termometro
    ws3 = wb.create_sheet(title="Calibracion_Termometro")
    ws3.append(["Temp_Indicada_C", "Temp_Patron_C", "Incertidumbre_Expandida_U_C", "Factor_k"])
    puntos_demo = [
        [0.00, 0.02, 0.04, 2.00],
        [10.00, 10.03, 0.04, 2.00],
        [20.00, 20.04, 0.05, 2.00],
        [30.00, 30.05, 0.05, 2.00],
        [40.00, 40.07, 0.06, 2.00],
        [50.00, 50.08, 0.06, 2.00],
        [90.00, 90.09, 0.07, 2.00],
    ]
    for p in puntos_demo:
        ws3.append(p)

    wb.save(filepath)
    print(f"Plantilla Excel creada exitosamente en '{filepath}'.")
    return filepath

def leer_datos_excel(filepath_or_stream) -> Dict:
    """
    Extrae los datos de temperatura y calibración de CUALQUIER archivo Excel,
    ignorando rigurosamente las filas de metadatos (Resolución 0.01, Deriva 0.05, Homogeneidad 0.1).
    """
    lecturas = []
    resolucion = 0.01
    deriva = 0.05
    homogeneidad = 0.10
    puntos_cal = []

    keywords_meta = ['resolucion', 'resolución', 'deriva', 'homogeneidad', 'gradiente', 'parametro', 'parámetro', 'codigo', 'código', 'unidad', 'notas']
    keywords_temp = ['temperatura', 'temp', 'concreto', 'hormigon', 'hormigón', 'c°', '°c', 'grados', 'grado', 'medicion', 'medición']
    keywords_ignore_cols = ['lectura_no', 'no', 'num', 'n°', 'item', 'id', 'pos', '#', 'indice', 'índice', 'renglon']

    with pd.ExcelFile(filepath_or_stream) as xls:
        hojas = xls.sheet_names

        # -------------------------------------------------------------
        # PASO 1: EXTRAER METADATOS Y EXTRAER SU UBICACIÓN DE FILAS
        # -------------------------------------------------------------
        filas_metadatos = set()
        for sheet in hojas:
            try:
                df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)
                for idx_row in range(len(df_raw)):
                    row_cells = [str(cell).strip() for cell in df_raw.iloc[idx_row].values if pd.notnull(cell)]
                    row_text = " ".join(row_cells).lower()

                    # Si la fila contiene palabras de metadatos, registrarla como fila de metadatos
                    if any(m in row_text for m in keywords_meta):
                        filas_metadatos.add(idx_row)

                    if 'resolucion' in row_text or 'resolución' in row_text:
                        nums = [float(x) for x in row_cells if is_float(x)]
                        if nums: resolucion = nums[0]
                    elif 'deriva' in row_text:
                        nums = [float(x) for x in row_cells if is_float(x)]
                        if nums: deriva = nums[0]
                    elif 'homogeneidad' in row_text or 'gradiente' in row_text:
                        nums = [float(x) for x in row_cells if is_float(x)]
                        if nums: homogeneidad = nums[0]
            except Exception:
                pass

        # -------------------------------------------------------------
        # PASO 2: EXTRAER TEMPERATURAS REALES IGNORANDO FILAS DE METADATOS
        # -------------------------------------------------------------
        for sheet in hojas:
            sheet_lower = str(sheet).lower()
            if any(k in sheet_lower for k in ['calib', 'cert', 'patron', 'patrón', 'param']) and len(hojas) > 1:
                continue

            try:
                # Leer hoja completa sin interpretar encabezado fijo para controlar índices de filas
                df_raw = pd.read_excel(xls, sheet_name=sheet, header=None)
                if df_raw.empty: continue

                # Buscar la fila que actúa como encabezado de la tabla de temperaturas (ej. 'Temperatura_Concreto_C' o 'Temperatura')
                header_row_idx = None
                target_col_idx = None

                for r in range(len(df_raw)):
                    row_vals = [str(v).strip().lower() for v in df_raw.iloc[r].values if pd.notnull(v)]
                    for c_idx, cell_val in enumerate(df_raw.iloc[r].values):
                        val_str = str(cell_val).strip().lower()
                        if any(kw in val_str for kw in keywords_temp) and not any(ign in val_str for ign in keywords_ignore_cols):
                            header_row_idx = r
                            target_col_idx = c_idx
                            break
                    if header_row_idx is not None:
                        break

                candidatos_vals = []

                if target_col_idx is not None and header_row_idx is not None:
                    # Extraer ÚNICAMENTE las filas posteriores al encabezado de la tabla
                    for r in range(header_row_idx + 1, len(df_raw)):
                        if r in filas_metadatos:
                            continue
                        val = df_raw.iloc[r, target_col_idx]
                        if pd.notnull(val) and is_float(str(val)):
                            candidatos_vals.append(float(val))

                # Si no encontramos encabezado explícito, escanear columnas numéricas ignorando filas de metadatos
                if not candidatos_vals:
                    for c_idx in range(df_raw.shape[1]):
                        col_cells = df_raw.iloc[:, c_idx]
                        header_name = str(col_cells.iloc[0]).strip().lower() if len(col_cells)>0 else ""
                        if any(ign in header_name for ign in keywords_ignore_cols) or any(m in header_name for m in keywords_meta):
                            continue

                        vals_col = []
                        for r_idx, val in enumerate(col_cells):
                            if r_idx in filas_metadatos:
                                continue
                            if pd.notnull(val) and is_float(str(val)):
                                vals_col.append(float(val))

                        if len(vals_col) >= 1:
                            # Filtrar valores diminutos de metadatos si están al inicio y difieren drásticamente de la media
                            if len(vals_col) >= 3:
                                median_val = np.median(vals_col)
                                # Si hay lecturas normales (ej. > 5 °C), descartar diminutivos (ej. 0.01, 0.05, 0.1) que están a la izquierda
                                vals_col = [v for v in vals_col if not (v < 1.0 and median_val > 5.0)]

                            if len(vals_col) > 0:
                                candidatos_vals = vals_col
                                break

                if candidatos_vals:
                    # Aplicar filtro final de remoción de metadatos preambulares (< 1.0 °C cuando las demás son temperaturas de concreto > 5 °C)
                    if len(candidatos_vals) >= 3:
                        mediana = np.median(candidatos_vals)
                        if mediana > 5.0:
                            candidatos_vals = [v for v in candidatos_vals if v >= 1.0]

                    lecturas = candidatos_vals
                    break

            except Exception as e:
                print(f"Error al analizar hoja {sheet}: {e}")

        # -------------------------------------------------------------
        # PASO 3: EXTRAER PUNTOS DE CALIBRACIÓN DEL CERTIFICADO
        # -------------------------------------------------------------
        for sheet in hojas:
            sheet_lower = str(sheet).lower()
            es_hoja_calib = any(k in sheet_lower for k in ['calib', 'cert', 'patron', 'patrón'])

            try:
                df_cal = pd.read_excel(xls, sheet_name=sheet)
                cols_str = " ".join([str(c).lower() for c in df_cal.columns])
                tiene_cols_calib = any(k in cols_str for k in ['indicada', 'temp_indicada']) or (any(k in cols_str for k in ['patron', 'patrón', 'ref']) and any(k in cols_str for k in ['u_', 'incertidumbre', 'expandida']))

                if (es_hoja_calib or tiene_cols_calib) and len(df_cal) >= 2 and len(df_cal.columns) >= 2:
                    col_ind = next((c for c in df_cal.columns if any(k in str(c).lower() for k in ['indicada', 'temp_ind', 'ind'])), df_cal.columns[0])
                    col_pat = next((c for c in df_cal.columns if any(k in str(c).lower() for k in ['patron', 'patrón', 'ref', 'temp_pat'])), df_cal.columns[1] if len(df_cal.columns)>1 else col_ind)
                    col_u = next((c for c in df_cal.columns if any(k in str(c).lower() for k in ['u_', 'incertidumbre', 'expandida'])), None)
                    col_k = next((c for c in df_cal.columns if any(k in str(c).lower() for k in ['k', 'factor'])), None)

                    puntos_temp = []
                    for _, r in df_cal.iterrows():
                        try:
                            t_ind = float(r[col_ind])
                            t_pat = float(r[col_pat]) if col_pat else t_ind
                            u_val = float(r[col_u]) if col_u and pd.notnull(r[col_u]) else 0.05
                            k_val = float(r[col_k]) if col_k and pd.notnull(r[col_k]) else 2.0
                            if -50 <= t_ind <= 300:
                                puntos_temp.append({
                                    "temp_indicada": t_ind,
                                    "temp_patron": t_pat,
                                    "correccion": t_pat - t_ind,
                                    "u_expandida": u_val,
                                    "factor_k": k_val
                                })
                        except Exception:
                            continue
                    if len(puntos_temp) >= 2:
                        puntos_cal = puntos_temp
                        break
            except Exception:
                pass

    return {
        "lecturas": lecturas,
        "resolucion": resolucion,
        "deriva": deriva,
        "homogeneidad": homogeneidad,
        "puntos_calibracion": puntos_cal
    }

def is_float(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False

if __name__ == "__main__":
    fp = crear_plantilla_excel()
    datos = leer_datos_excel(fp)
    print("Datos extraídos de Excel:", datos)
